"""
Excel import service.

Supported sheets:
  - "Projects"  → creates Project rows
  - "Requests"  → finds/creates Project+Unit, creates Request rows

Projects columns (case-insensitive):
  Name* | Client Name | Location | Status | Start Date | End Date

Requests columns (case-insensitive):
  Project Name* | Unit Name* | Unit Type | Title* | Category | Priority |
  Description | Supplier | Expected Delivery
"""
import uuid
from datetime import date, datetime
from io import BytesIO
from typing import Any

import openpyxl
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.project import Project, ProjectStatus
from app.models.request import Request, RequestCategory, RequestPriority, RequestStatus
from app.models.unit import Unit, UnitType


# ── helpers ──────────────────────────────────────────────────────────────────

def _header_map(sheet) -> dict[str, int]:
    """Return {normalized_col_name: col_index} from first row."""
    row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    return {str(c).strip().lower(): i for i, c in enumerate(row) if c}


def _val(row: tuple, headers: dict[str, int], key: str) -> Any:
    idx = headers.get(key.lower())
    if idx is None:
        return None
    v = row[idx]
    return v.strip() if isinstance(v, str) else v


def _parse_date(v) -> date | None:
    if v is None:
        return None
    if isinstance(v, (date, datetime)):
        return v.date() if isinstance(v, datetime) else v
    try:
        return datetime.strptime(str(v).strip(), "%Y-%m-%d").date()
    except ValueError:
        return None


def _parse_enum(enum_cls, v: str | None, default=None):
    if v is None:
        return default
    try:
        return enum_cls[v.strip().upper()]
    except KeyError:
        return default


# ── main import function ──────────────────────────────────────────────────────

async def import_excel(
    db: AsyncSession, file_bytes: bytes, user_id: uuid.UUID
) -> dict:
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    summary = {"projects_created": 0, "requests_created": 0, "errors": []}

    # ── Projects sheet ────────────────────────────────────────────────────────
    if "Projects" in wb.sheetnames:
        ws = wb["Projects"]
        headers = _header_map(ws)
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = _val(row, headers, "name")
            if not name:
                continue
            try:
                project = Project(
                    name=str(name).strip(),
                    client_name=_val(row, headers, "client name"),
                    location=_val(row, headers, "location"),
                    status=_parse_enum(
                        ProjectStatus,
                        _val(row, headers, "status"),
                        ProjectStatus.PLANNING,
                    ),
                    start_date=_parse_date(_val(row, headers, "start date")),
                    end_date=_parse_date(_val(row, headers, "end date")),
                    created_by=user_id,
                )
                db.add(project)
                summary["projects_created"] += 1
            except Exception as e:
                summary["errors"].append(f"Project row '{name}': {e}")

        await db.flush()

    # ── Requests sheet ────────────────────────────────────────────────────────
    if "Requests" in wb.sheetnames:
        ws = wb["Requests"]
        headers = _header_map(ws)

        # Cache project lookups to avoid N+1
        project_cache: dict[str, Project] = {}
        unit_cache: dict[tuple[str, str], Unit] = {}

        for row in ws.iter_rows(min_row=2, values_only=True):
            project_name = _val(row, headers, "project name")
            unit_name = _val(row, headers, "unit name")
            title = _val(row, headers, "title")
            if not all([project_name, unit_name, title]):
                continue

            project_name = str(project_name).strip()
            unit_name = str(unit_name).strip()
            title = str(title).strip()

            try:
                # Find project
                if project_name not in project_cache:
                    result = await db.execute(
                        select(Project).where(Project.name == project_name)
                    )
                    proj = result.scalar_one_or_none()
                    if not proj:
                        summary["errors"].append(
                            f"Request '{title}': project '{project_name}' not found"
                        )
                        continue
                    project_cache[project_name] = proj
                project = project_cache[project_name]

                # Find or create unit
                cache_key = (project_name, unit_name)
                if cache_key not in unit_cache:
                    result = await db.execute(
                        select(Unit).where(
                            Unit.project_id == project.id, Unit.name == unit_name
                        )
                    )
                    unit = result.scalar_one_or_none()
                    if not unit:
                        raw_type = _val(row, headers, "unit type")
                        unit = Unit(
                            project_id=project.id,
                            name=unit_name,
                            type=_parse_enum(UnitType, raw_type, UnitType.APARTMENT),
                        )
                        db.add(unit)
                        await db.flush()
                    unit_cache[cache_key] = unit
                unit = unit_cache[cache_key]

                request = Request(
                    unit_id=unit.id,
                    title=title,
                    description=_val(row, headers, "description"),
                    category=_parse_enum(
                        RequestCategory,
                        _val(row, headers, "category"),
                        RequestCategory.OTHER,
                    ),
                    priority=_parse_enum(
                        RequestPriority,
                        _val(row, headers, "priority"),
                        RequestPriority.MEDIUM,
                    ),
                    status=RequestStatus.MATERIAL_REQUEST,
                    supplier_name=_val(row, headers, "supplier"),
                    expected_delivery_date=_parse_date(
                        _val(row, headers, "expected delivery")
                    ),
                    created_by=user_id,
                )
                db.add(request)
                summary["requests_created"] += 1
            except Exception as e:
                summary["errors"].append(f"Request row '{title}': {e}")

        await db.flush()

    await db.commit()
    wb.close()
    return summary
